from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter

from app.config import load_config
from app.services.threshold_engine import detect_metric_type, status_for, threshold_for


KPI_HEADERS = {
    "kpi",
    "kpi name",
    "metric",
    "metric name",
    "parameter",
    "name",
    "description",
    "kpi description",
}
CATEGORY_HEADERS = {"category", "domain", "type", "group"}
REMARK_HEADERS = {"remark", "remarks", "comment", "comments"}
IGNORED_VALUE_HEADERS = KPI_HEADERS | CATEGORY_HEADERS | REMARK_HEADERS | {
    "status",
    "threshold",
    "target",
    "sr",
    "s.no",
    "sno",
}


@dataclass(frozen=True)
class ParsedKpiRecord:
    sheet_name: str
    row_number: int
    kpi_name: str
    category: str | None
    week_label: str
    value_text: str | None
    value_number: float | None
    value_cell: str
    remarks_cell: str | None
    current_remark: str | None
    threshold: float
    status: str
    metric_type: str


def normalize_header(value: Any) -> str:
    return str(value or "").strip().lower()


def cell_to_text(cell: Cell) -> str | None:
    if cell.value is None:
        return None
    return str(cell.value).strip()


def number_from_cell(cell: Cell) -> float | None:
    value = cell.value
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip().replace("%", "").replace(",", "")
    try:
        return float(text)
    except ValueError:
        return None


def find_header_row(sheet) -> tuple[int, dict[str, int]] | None:
    for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 20)):
        headers = {normalize_header(cell.value): cell.column for cell in row if normalize_header(cell.value)}
        if headers.keys() & KPI_HEADERS:
            return row[0].row, headers
    return None


def first_matching_column(headers: dict[str, int], candidates: set[str]) -> int | None:
    for header, column in headers.items():
        if header in candidates or any(candidate in header for candidate in candidates):
            return column
    return None


def value_columns(headers: dict[str, int]) -> list[tuple[str, int]]:
    columns: list[tuple[str, int]] = []
    for header, column in headers.items():
        if header not in IGNORED_VALUE_HEADERS and not any(ignored in header for ignored in REMARK_HEADERS):
            columns.append((header.upper() if header.startswith("wk") else header.title(), column))
    return columns


def parse_workbook(path: Path) -> list[ParsedKpiRecord]:
    config = load_config()
    workbook = load_workbook(path, data_only=True)
    parsed: list[ParsedKpiRecord] = []

    for sheet in workbook.worksheets:
        header_info = find_header_row(sheet)
        if not header_info:
            continue

        header_row, headers = header_info
        kpi_col = first_matching_column(headers, KPI_HEADERS)
        if not kpi_col:
            continue

        category_col = first_matching_column(headers, CATEGORY_HEADERS)
        remarks_col = first_matching_column(headers, REMARK_HEADERS)
        week_columns = value_columns(headers)
        current_category: str | None = None

        for row_number in range(header_row + 1, sheet.max_row + 1):
            kpi_cell = sheet.cell(row=row_number, column=kpi_col)
            kpi_name = cell_to_text(kpi_cell)
            if not kpi_name:
                continue

            category = cell_to_text(sheet.cell(row=row_number, column=category_col)) if category_col else None
            if category:
                current_category = category
            else:
                category = current_category

            remarks_cell = sheet.cell(row=row_number, column=remarks_col) if remarks_col else None
            remark_text = cell_to_text(remarks_cell) if remarks_cell else None

            for week_label, column in week_columns:
                value_cell = sheet.cell(row=row_number, column=column)
                value_text = cell_to_text(value_cell)
                value_number = number_from_cell(value_cell)
                if value_text is None and value_number is None:
                    continue

                metric_type = detect_metric_type(kpi_name, category)
                threshold = threshold_for(kpi_name, metric_type, config)
                status = status_for(value_number, threshold, metric_type)
                value_address = f"{get_column_letter(column)}{row_number}"
                remark_address = f"{get_column_letter(remarks_col)}{row_number}" if remarks_col else None

                parsed.append(
                    ParsedKpiRecord(
                        sheet_name=sheet.title,
                        row_number=row_number,
                        kpi_name=kpi_name,
                        category=category,
                        week_label=week_label,
                        value_text=value_text,
                        value_number=value_number,
                        value_cell=value_address,
                        remarks_cell=remark_address,
                        current_remark=remark_text,
                        threshold=threshold,
                        status=status,
                        metric_type=metric_type,
                    )
                )

    workbook.close()
    return parsed
